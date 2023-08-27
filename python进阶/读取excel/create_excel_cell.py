import openpyxl

book = openpyxl.Workbook()

sh = book.active

sh.title = "工资表"

book.save('信息.xlsx')

sh1 = book.create_sheet('年龄表_最后')
sh2 = book.create_sheet('年龄表_最前',0)
sh3 = book.create_sheet('年龄表2',1)


sh = book['工资表']
sh['A1'] = '你好'

print(sh['A1'].value)

sh.cell(2,2).value = '白月黑羽'

print(sh.cell(1,1).value)
book.save('信息.xlsx')


