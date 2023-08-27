import openpyxl
from openpyxl.styles import Font, colors, PatternFill

wb = openpyxl.load_workbook('./income.xlsx')

sheet1 = wb['2017']
sheet1['A1'].font = Font(
    color=colors.BLUE,
    size=15,
    bold=True,
    italic=True

)
sheet1['A1'].fill = PatternFill("solid","E39191")
fill = PatternFill("solid","E39191")

sheet1['B1'].font = Font(color='981818')
font = Font(color='981818')
for y in range(1,100):
    sheet1.cell(row=3,column=y).font = font

for x in range(1,100):
    sheet1.cell(row=x, column=2).font = font

sheet2 = wb['2018']
sheet2.insert_rows(2)
sheet2.insert_rows(3, 3)
sheet2.insert_cols(2)
sheet2.insert_cols(2, 3)

wb.save('income-1.xlsx')
